"""
excel_export.py
Modul tambahan (Presentation-support Layer) - bertugas mengekspor hasil
ranking TOPSIS ke file Excel (.xlsx) dengan formatting profesional
menggunakan openpyxl (header berwarna, auto-width kolom, border, dsb).

Modul ini TIDAK menyentuh logika perhitungan TOPSIS sama sekali -> hanya
mengubah bentuk output dari CSV menjadi Excel yang lebih rapi & profesional.
"""

import io
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# Palet warna warm cafe untuk styling Excel
WARNA_COKLAT = "6F4E37"
WARNA_CREAM = "F8F4EC"
WARNA_SAGE = "A3B18A"
WARNA_TEKS_TERANG = "FFFFFF"


def _style_header(ws, jumlah_kolom: int, baris: int = 1) -> None:
    """Memberi styling header: background coklat, teks putih tebal, border."""
    fill = PatternFill(start_color=WARNA_COKLAT, end_color=WARNA_COKLAT, fill_type="solid")
    font = Font(bold=True, color=WARNA_TEKS_TERANG, size=11, name="Calibri")
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin", color="4A3423"),
        right=Side(style="thin", color="4A3423"),
        top=Side(style="thin", color="4A3423"),
        bottom=Side(style="thin", color="4A3423"),
    )
    for col_idx in range(1, jumlah_kolom + 1):
        cell = ws.cell(row=baris, column=col_idx)
        cell.fill = fill
        cell.font = font
        cell.alignment = align
        cell.border = border
    ws.row_dimensions[baris].height = 28


def _style_data_rows(ws, baris_awal: int, baris_akhir: int, jumlah_kolom: int, kolom_ranking: int = 1) -> None:
    """Memberi styling baris data: border tipis, selang-seling warna cream/putih,
    highlight khusus untuk 3 ranking teratas dengan nuansa sage."""
    border = Border(
        left=Side(style="thin", color="D9D0C3"),
        right=Side(style="thin", color="D9D0C3"),
        top=Side(style="thin", color="D9D0C3"),
        bottom=Side(style="thin", color="D9D0C3"),
    )
    fill_cream = PatternFill(start_color=WARNA_CREAM, end_color=WARNA_CREAM, fill_type="solid")
    fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    fill_sage = PatternFill(start_color="E3E9D8", end_color="E3E9D8", fill_type="solid")

    for row_idx in range(baris_awal, baris_akhir + 1):
        ranking_val = ws.cell(row=row_idx, column=kolom_ranking).value
        is_top3 = isinstance(ranking_val, (int, float)) and ranking_val <= 3

        if is_top3:
            row_fill = fill_sage
        elif (row_idx - baris_awal) % 2 == 0:
            row_fill = fill_white
        else:
            row_fill = fill_cream

        for col_idx in range(1, jumlah_kolom + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = row_fill
            cell.border = border
            cell.font = Font(
                color="4A3423",
                bold=is_top3,
                size=10.5,
                name="Calibri",
            )
            if col_idx == kolom_ranking:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")


def _auto_width(ws, df: pd.DataFrame, baris_header: int = 1) -> None:
    """Menghitung lebar kolom otomatis berdasarkan konten terpanjang."""
    for idx, kolom in enumerate(df.columns, start=1):
        panjang_header = len(str(kolom))
        try:
            panjang_data = df[kolom].astype(str).map(len).max()
        except (ValueError, TypeError):
            panjang_data = 10
        panjang_maks = max(panjang_header, int(panjang_data) if pd.notna(panjang_data) else 10)
        lebar = min(max(panjang_maks + 4, 12), 45)
        ws.column_dimensions[get_column_letter(idx)].width = lebar


def buat_excel_hasil_ranking(
    df_tampil: pd.DataFrame,
    nama_aplikasi: str = "YumRank — DSS Pemilihan Tempat Makan",
    metode: str = "TOPSIS",
    bobot_info: dict | None = None,
) -> bytes:
    """
    Membuat file Excel (.xlsx) profesional berisi hasil ranking TOPSIS.

    Sheet 1 "Hasil Ranking": tabel hasil dengan header berwarna, border,
    auto-width, highlight top-3, dan format angka.
    Sheet 2 "Ringkasan": info aplikasi, metode, tanggal export, dan bobot
    kriteria yang digunakan (jika disediakan).

    Args:
        df_tampil: DataFrame hasil ranking yang siap ditampilkan/diekspor
                    (kolom sudah dalam bentuk label akhir, TIDAK diubah
                    nilainya sama sekali -- hanya diformat visualnya).
        nama_aplikasi: nama aplikasi untuk judul sheet ringkasan.
        metode: nama metode DSS yang dipakai.
        bobot_info: dict {label_kriteria: persen} opsional untuk dicatat
                    di sheet ringkasan.

    Returns:
        bytes: konten file .xlsx siap diunduh lewat st.download_button.
    """
    wb = Workbook()

    # -------------------- SHEET 1: HASIL RANKING --------------------
    ws = wb.active
    ws.title = "Hasil Ranking"

    jumlah_kolom = len(df_tampil.columns)

    # Judul di baris paling atas
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=jumlah_kolom)
    judul_cell = ws.cell(row=1, column=1, value=f"☕ {nama_aplikasi} — Hasil Ranking ({metode})")
    judul_cell.font = Font(bold=True, size=14, color=WARNA_COKLAT, name="Calibri")
    judul_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=jumlah_kolom)
    sub_cell = ws.cell(
        row=2, column=1,
        value=f"Diekspor pada {datetime.now().strftime('%d %B %Y, %H:%M')}"
    )
    sub_cell.font = Font(italic=True, size=9, color="8A7968", name="Calibri")
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    baris_header = 4
    # Tulis header kolom
    for col_idx, kolom in enumerate(df_tampil.columns, start=1):
        ws.cell(row=baris_header, column=col_idx, value=str(kolom))
    _style_header(ws, jumlah_kolom, baris=baris_header)

    # Tulis data
    baris_data_awal = baris_header + 1
    for r_offset, (_, row) in enumerate(df_tampil.iterrows()):
        for c_idx, kolom in enumerate(df_tampil.columns, start=1):
            nilai = row[kolom]
            if isinstance(nilai, float):
                nilai = round(float(nilai), 4)
            ws.cell(row=baris_data_awal + r_offset, column=c_idx, value=nilai)
    baris_data_akhir = baris_data_awal + len(df_tampil) - 1

    kolom_ranking_idx = list(df_tampil.columns).index("Ranking") + 1 if "Ranking" in df_tampil.columns else 1
    _style_data_rows(ws, baris_data_awal, baris_data_akhir, jumlah_kolom, kolom_ranking=kolom_ranking_idx)
    _auto_width(ws, df_tampil, baris_header=baris_header)

    ws.freeze_panes = ws.cell(row=baris_data_awal, column=1)

    # Format angka desimal 4 digit untuk kolom skor/jarak
    for col_idx, kolom in enumerate(df_tampil.columns, start=1):
        if any(kata in str(kolom).lower() for kata in ["skor", "jarak"]):
            for row_idx in range(baris_data_awal, baris_data_akhir + 1):
                ws.cell(row=row_idx, column=col_idx).number_format = "0.0000"

    # -------------------- SHEET 2: RINGKASAN --------------------
    ws2 = wb.create_sheet("Ringkasan")
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 40

    ws2.merge_cells("A1:B1")
    c = ws2.cell(row=1, column=1, value=f"☕ {nama_aplikasi}")
    c.font = Font(bold=True, size=14, color=WARNA_COKLAT, name="Calibri")
    ws2.row_dimensions[1].height = 26

    info_rows = [
        ("Metode DSS", metode),
        ("Tanggal Export", datetime.now().strftime("%d %B %Y, %H:%M")),
        ("Jumlah Alternatif", str(len(df_tampil))),
    ]
    baris = 3
    for label, nilai in info_rows:
        lc = ws2.cell(row=baris, column=1, value=label)
        vc = ws2.cell(row=baris, column=2, value=nilai)
        lc.font = Font(bold=True, color=WARNA_COKLAT, size=10.5, name="Calibri")
        vc.font = Font(color="4A3423", size=10.5, name="Calibri")
        lc.fill = PatternFill(start_color=WARNA_CREAM, end_color=WARNA_CREAM, fill_type="solid")
        vc.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        baris += 1

    if bobot_info:
        baris += 1
        ws2.merge_cells(start_row=baris, start_column=1, end_row=baris, end_column=2)
        judul_bobot = ws2.cell(row=baris, column=1, value="Bobot Kriteria yang Digunakan")
        judul_bobot.font = Font(bold=True, size=11, color=WARNA_TEKS_TERANG, name="Calibri")
        judul_bobot.fill = PatternFill(start_color=WARNA_SAGE, end_color=WARNA_SAGE, fill_type="solid")
        judul_bobot.alignment = Alignment(horizontal="center")
        baris += 1
        for label, persen in bobot_info.items():
            lc = ws2.cell(row=baris, column=1, value=label)
            vc = ws2.cell(row=baris, column=2, value=f"{persen}%")
            lc.font = Font(color="4A3423", size=10, name="Calibri")
            vc.font = Font(color="4A3423", size=10, name="Calibri")
            fill = PatternFill(
                start_color=WARNA_CREAM if baris % 2 == 0 else "FFFFFF",
                end_color=WARNA_CREAM if baris % 2 == 0 else "FFFFFF",
                fill_type="solid",
            )
            lc.fill = fill
            vc.fill = fill
            baris += 1

    # Simpan ke bytes buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
